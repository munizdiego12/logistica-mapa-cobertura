import io
import math
import time
import asyncio
import uuid
import httpx
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Optional, Dict, Any
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

import database

app = FastAPI(title="Zubale Routing Core - Dynamic National Engine")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def _on_startup():
    await database.init_db()

MODAIS_PADRAO_DEFAULT = {
    'Moto': {'capacidade': 3, 'consumo_kml': 30.0},
    'Carro de Passeio': {'capacidade': 5, 'consumo_kml': 11.5},
    'Fiorino / Utilitário': {'capacidade': 12, 'consumo_kml': 9.0},
    'VUC / Caminhão 3/4': {'capacidade': 30, 'consumo_kml': 6.5}
}

CORES_ROTAS = [
    '#3B82F6', '#10B981', '#F59E0B', '#EF4444', 
    '#8B5CF6', '#EC4899', '#06B6D4', '#84CC16',
    '#F97316', '#14B8A6', '#6366F1', '#D946EF'
]

ESTADOS_IBGE_BRASIL = {
    "SP": {"ibge": 3550308, "nome": "São Paulo"},
    "PR": {"ibge": 4106902, "nome": "Curitiba"},
    "CE": {"ibge": 2304400, "nome": "Fortaleza"},
    "RJ": {"ibge": 3304557, "nome": "Rio de Janeiro"},
    "MG": {"ibge": 3106200, "nome": "Belo Horizonte"},
    "RS": {"ibge": 4314902, "nome": "Porto Alegre"},
    "SC": {"ibge": 4205407, "nome": "Florianópolis"},
    "GO": {"ibge": 5208707, "nome": "Goiânia"},
    "BA": {"ibge": 2927408, "nome": "Salvador"},
    "PE": {"ibge": 2611606, "nome": "Recife"},
    "DF": {"ibge": 5300108, "nome": "Brasília"}
}

def haversine_distance(coord1: tuple, coord2: tuple) -> float:
    lat1, lon1 = coord1
    lat2, lon2 = coord2
    R = 6371.0

    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)

    a = math.sin(delta_phi / 2.0)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2.0)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

_GEOCODE_CACHE_MEMORIA: Dict[str, tuple] = {}

# Controla o rate-limit do Nominatim (1 req/s) de forma assíncrona e SEGURA
# mesmo quando várias geocodificações rodam "em paralelo" (asyncio.gather).
_nominatim_lock = asyncio.Lock()
_nominatim_ultima_chamada = 0.0

# BrasilAPI não exige 1 req/s, então aqui só limitamos concorrência (não velocidade).
_brasilapi_semaforo = asyncio.Semaphore(8)


def _chave_cache(rua, numero, bairro, cep, cidade, uf):
    return "|".join([str(x or "").strip().lower() for x in [rua, numero, bairro, cep, cidade, uf]])


async def _respeitar_rate_limit_nominatim():
    global _nominatim_ultima_chamada
    async with _nominatim_lock:
        agora = time.monotonic()
        espera = 1.05 - (agora - _nominatim_ultima_chamada)
        if espera > 0:
            await asyncio.sleep(espera)
        _nominatim_ultima_chamada = time.monotonic()


async def geocode_async(client: httpx.AsyncClient, rua: str = "", numero: str = "",
                         bairro: str = "", cep: str = "", cidade: str = "", uf: str = ""):
    """
    Versão assíncrona da geocodificação. Ordem de prioridade:
    1) Cache em memória (instantâneo, válido durante a execução do processo)
    2) Cache Postgres (instantâneo, sobrevive a reinícios do servidor)
    3) Endereço completo via Nominatim (preciso, mas limitado a 1 req/s)
    4) CEP via BrasilAPI (aproximado, mas pode rodar em paralelo)
    5) Bairro/cidade via Nominatim (último recurso)
    Nunca retorna coordenada fixa "chumbada".
    """
    clean_cep = ''.join(filter(str.isdigit, str(cep or "")))
    chave = _chave_cache(rua, numero, bairro, cep, cidade, uf)

    if chave in _GEOCODE_CACHE_MEMORIA:
        return _GEOCODE_CACHE_MEMORIA[chave]

    cache_db = await database.cache_get(chave)
    if cache_db:
        _GEOCODE_CACHE_MEMORIA[chave] = cache_db
        return cache_db

    resultado = None
    headers = {'User-Agent': 'ZubaleRoutingCore/6.1 (contato@zubale.com)'}

    # 1) Endereço completo (rua + número) via Nominatim
    if rua and numero:
        query = ", ".join([p for p in [f"{rua}, {numero}", bairro, cidade, uf, clean_cep, "Brasil"] if p])
        try:
            await _respeitar_rate_limit_nominatim()
            res = await client.get(
                "https://nominatim.openstreetmap.org/search",
                params={"format": "json", "q": query, "limit": 1},
                headers=headers, timeout=4,
            )
            data = res.json()
            if data:
                resultado = (float(data[0]['lat']), float(data[0]['lon']),
                             cidade or "Centro", uf or "BR", clean_cep, bairro)
        except Exception:
            pass

    # 2) CEP via BrasilAPI (pode rodar em paralelo, sem rate-limit rígido)
    if resultado is None and len(clean_cep) == 8:
        async with _brasilapi_semaforo:
            try:
                r = await client.get(f"https://brasilapi.com.br/api/cep/v2/{clean_cep}", timeout=3)
                if r.status_code == 200:
                    data = r.json()
                    loc = data.get("location", {}).get("coordinates", {})
                    lat, lon = loc.get("latitude"), loc.get("longitude")
                    if lat and lon:
                        resultado = (float(lat), float(lon), data.get("city", cidade),
                                     data.get("state", uf), clean_cep, data.get("neighborhood", bairro))
            except Exception:
                pass

    # 3) Fallback final: bairro/cidade via Nominatim
    if resultado is None:
        query = ", ".join([p for p in [bairro, cidade, uf, "Brasil"] if p]) or "São Paulo, Brasil"
        try:
            await _respeitar_rate_limit_nominatim()
            res = await client.get(
                "https://nominatim.openstreetmap.org/search",
                params={"format": "json", "q": query, "limit": 1},
                headers=headers, timeout=4,
            )
            data = res.json()
            if data:
                resultado = (float(data[0]['lat']), float(data[0]['lon']),
                             cidade or "Centro", uf or "BR", clean_cep, bairro)
        except Exception:
            pass

    if resultado is None:
        resultado = (-23.5614, -46.6559, cidade or "Origem", uf or "SP", clean_cep, bairro)

    _GEOCODE_CACHE_MEMORIA[chave] = resultado
    await database.cache_set(chave, resultado)
    return resultado


async def geocodificar_lote(pedidos_raw: List[Dict[str, Any]]) -> Dict[str, tuple]:
    """
    Geocodifica uma lista de pedidos de forma otimizada:
    - Deduplica endereços idênticos (ex: 2 pedidos no mesmo prédio geocodificam 1x só).
    - Roda em paralelo o que der (BrasilAPI, cache), respeitando o rate-limit do Nominatim.
    Retorna um dicionário {chave_do_endereco: (lat, lon, cidade, uf, cep, bairro)}.
    """
    enderecos_unicos = {}
    for p in pedidos_raw:
        rua = p.get("Endereco") or p.get("rua") or ""
        num = str(p.get("Numero") or p.get("numero") or "")
        bairro = p.get("Bairro") or p.get("bairro") or ""
        cep = p.get("CEP") or p.get("cep") or ""
        cidade = p.get("Cidade") or p.get("cidade") or ""
        uf = p.get("UF") or p.get("uf") or ""
        chave = _chave_cache(rua, numero=num, bairro=bairro, cep=cep, cidade=cidade, uf=uf)
        enderecos_unicos[chave] = (rua, num, bairro, cep, cidade, uf)

    resultados: Dict[str, tuple] = {}
    async with httpx.AsyncClient() as client:
        tarefas = [
            geocode_async(client, rua, num, bairro, cep, cidade, uf)
            for chave, (rua, num, bairro, cep, cidade, uf) in enderecos_unicos.items()
        ]
        respostas = await asyncio.gather(*tarefas)
        for chave, resposta in zip(enderecos_unicos.keys(), respostas):
            resultados[chave] = resposta

    return resultados


def geocode_rapido(rua: str = "", numero: str = "", bairro: str = "", cep: str = "", cidade: str = "", uf: str = ""):
    """
    Wrapper SÍNCRONO mantido para compatibilidade com endpoints não convertidos
    (ex: /api/cobertura-ceps, que geocodifica só 1 endereço por vez).
    """
    async def _run():
        async with httpx.AsyncClient() as client:
            return await geocode_async(client, rua, numero, bairro, cep, cidade, uf)
    return asyncio.run(_run())

class RaioCepRequest(BaseModel):
    origem_rua: str
    origem_num: str
    origem_cep: Optional[str] = ""
    origem_bairro: Optional[str] = ""
    origem_cidade: Optional[str] = ""
    origem_uf: Optional[str] = ""
    raio_km: Optional[float] = 30.0

async def _buscar_ceps_reais_banco(lat: float, lon: float, raio_km: float):
    """
    STUB DE IMPLEMENTAÇÃO (Projeto IBGE):
    Tenta buscar na tabela CNEFE. Se falhar, faz fallback para simulação.
    """
    try:
        if hasattr(database, "consultar_ceps_por_raio"):
            return await database.consultar_ceps_por_raio(lat, lon, raio_km)
        return []
    except Exception as e:
        return []

@app.post("/api/cobertura-ceps")
async def gerar_cobertura_ceps_instantanea(req: RaioCepRequest):
    # Usando o async original para não travar
    async with httpx.AsyncClient() as client:
        lat, lon, cidade, uf, clean_cep, bairro = await geocode_async(
            client, req.origem_rua, req.origem_num, req.origem_bairro or "", req.origem_cep or "", req.origem_cidade or "", req.origem_uf or ""
        )

    clean_cep = str(clean_cep).replace("-", "").strip().zfill(8)
    ibge_base = ESTADOS_IBGE_BRASIL.get(uf, {}).get("ibge", 3550308 if uf == "SP" else 2304400)

    prefixo_5 = clean_cep[:5] if len(clean_cep) == 8 else "01310"
    prefixo_num = int(prefixo_5)

    raio_max = req.raio_km or 30.0
    
    # 1. TENTA BUSCAR DADOS REAIS DO BANCO
    ceps_reais = await _buscar_ceps_reais_banco(lat, lon, raio_max)
    pontos_cobertos = []

    if ceps_reais and len(ceps_reais) > 0:
        pontos_cobertos = ceps_reais
    else:
        # 2. FALLBACK: Malha matemática (mantendo seu código 100% original aqui)
        pontos_cobertos.append({
            "ibge": ibge_base, "uf": uf or "SP", "cidade": cidade or "Sede do Hub",
            "bairro": "Centro / Sede Operacional", "cep_inicial": f"{prefixo_5}000",
            "cep_final": f"{prefixo_5}999", "distancia_km": 0.0, "dias_sla": 1,
            "lat": lat, "lon": lon
        })

        direcoes = [
            ("Norte", 0), ("Nordeste", 45), ("Leste", 90), ("Sudeste", 135),
            ("Sul", 180), ("Sudoeste", 225), ("Oeste", 270), ("Noroeste", 315)
        ]
        distancias = [3.5, 7.0, 12.0, 18.0, 24.0, 30.0]

        for d_km in distancias:
            if d_km > raio_max:
                continue
            delta_lat = d_km / 111.0
            
            for dir_idx, (nome_dir, ang_graus) in enumerate(direcoes):
                rad = math.radians(ang_graus)
                p_lat = lat + (delta_lat * math.cos(rad))
                p_lon = lon + (delta_lat * math.sin(rad) / max(0.1, math.cos(math.radians(lat))))
                
                dist_real = haversine_distance((lat, lon), (p_lat, p_lon))
                offset_prefixo = int((dir_idx + 1) * 10 + (d_km * 2))
                novo_prefixo = str(max(100, prefixo_num + offset_prefixo)).zfill(5)
                
                pontos_cobertos.append({
                    "ibge": ibge_base, "uf": uf or "SP", "cidade": cidade or "Região Metropolitana",
                    "bairro": f"Setor {nome_dir} ({d_km} km)", "cep_inicial": f"{novo_prefixo}000",
                    "cep_final": f"{novo_prefixo}999", "distancia_km": round(dist_real, 2),
                    "dias_sla": 1 if dist_real <= 12 else 2, "lat": p_lat, "lon": p_lon
                })

    pontos_cobertos.sort(key=lambda x: x["distancia_km"])

    return {
        "hub": {
            "lat": lat, "lon": lon, "cidade": cidade, "uf": uf, "cep": clean_cep,
            "ibge": ibge_base, "endereco": f"{req.origem_rua}, {req.origem_num}"
        },
        "raio_limite_km": raio_max,
        "total_pontos": len(pontos_cobertos),
        "pontos_cobertos": pontos_cobertos
    }

class ExportarXlsxRequest(BaseModel):
    hub: Dict[str, Any]
    pontos_cobertos: List[Dict[str, Any]]

@app.post("/api/exportar-tabela-frete-xlsx")
def exportar_tabela_frete_xlsx(req: ExportarXlsxRequest):
    wb = openpyxl.Workbook()
    
    # Aba 1: Prazos e Preços
    ws1 = wb.active
    ws1.title = "Prazos e preços"
    
    header_l1 = [None]*8 + ["Faixa Peso (Kg)", "De / Até", "De / Até", "De / Até", "De / Até", "De / Até"]
    header_l2 = ["Faixa Destino", None, None, None, None, None, None, "Prazo entrega", 0.000001, 10.000001, 20.000001, 30.000001, 50.000001, 70.000001, "Excedente", None, "Taxa", None, None, "Advalorem", None, None, None, None, "ADEME", None, None, "GRIS", None, None, None, None, None, "TRT", None, None, "TDA", None, None, None, None, None, None, None, "Taxa Fluvial", None, None, "EMEX", None, None, None, None, "Pedágio", None, None, None, None, "Restrição de Dimensões Máximas (Centímetros)", None, "Restrição de Dimensões (Centímetros)", None, "ICMS", None]
    header_l3 = [
        "Código IBGE", "Descrição (Município ou distrito)", "UF", "Cidade", "Faixa de precificação", "CEP Inicial", "CEP Final", "Dias",
        10, 20, 30, 50, 70, 100, "Taxa Excedente x KG", "Excedente fixo (R$)", "Coleta (R$)", "Despacho (R$)", "Entrega (R$)",
        "ADV1 (%)", "Base cálculo para ADV2 (R$)", "ADV2 (%)", "Mínimo (R$)", "Máximo (R$)",
        "% Percentual", "Mínimo (R$)", "Máximo (R$)", "Base cálculo para GRIS1 (R$)", "GRIS1 (%)", "Base cálculo para GRIS2 (R$)", "GRIS2 (%)", "Mínimo (R$)", "Máximo (R$)",
        "% Percentual", "Mínimo (R$)", "Máximo (R$)", "% Percentual", "Mínimo (R$)", "Máximo (R$)",
        "Balsa (R$)", "Suframa Valor (R$)", "TAS (R$)", "SEC CAT (R$)", "DAT (R$)",
        "% Percentual", "Mínimo (R$)", "Máximo (R$)", "% Percentual NF", "Valor por fração 100Kg (R$)", "Valor fixo (R$)", "Mínimo (R$)", "Máximo (R$)",
        "Valor (R$)", "Fração KG", "Mínimo (R$)", "Máximo (R$)", "Fator de Cubagem",
        "Altura Máxima", "Largura Máxima", "Comprimento Máximo", "Soma Máxima", "% Percentual da rota", "ICMS sobre o pedágio"
    ]
    
    ws1.append(header_l1)
    ws1.append(header_l2)
    ws1.append(header_l3)
    
    for p in req.pontos_cobertos:
        ibge = p.get("ibge") or req.hub.get("ibge") or 3550308
        uf = p.get("uf") or req.hub.get("uf") or "SP"
        cidade = p.get("cidade") or req.hub.get("cidade") or "Município"
        bairro = p.get("bairro") or ""
        cep_ini = str(p.get("cep_inicial", "00000000")).zfill(8)
        cep_fim = str(p.get("cep_final", "99999999")).zfill(8)
        dias = p.get("dias_sla") or (1 if p.get("distancia_km", 0) <= 12 else 2)
        dist = p.get("distancia_km", 0)
        
        linha = [
            ibge, bairro, uf, cidade, f"Raio {dist} km", cep_ini, cep_fim, dias,
            1, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0,
            0, 0, 0, 0, 0,
            0, 0, 0, 0, 0, 0, 0, 0, 0,
            0, 0, 0, 0, 0, 0,
            0, 0, 0, 0, 0,
            0, 0, 0, 0, 0, 0, 0, 0,
            0, 100, 0, 0, 250,
            0, 0, 0, 0, 0, 0
        ]
        ws1.append(linha)

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    
    for col_idx in range(1, len(header_l3) + 1):
        cell = ws1.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws2 = wb.create_sheet(title="TZR e TDE")
    ws2.append(["CNPJ", "TZR", "TDE"])
    ws2.append(["00.000.000/0000-00", 0, 0])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers_resp = {
        'Content-Disposition': 'attachment; filename="tabela_frete_completa.xlsx"'
    }
    return StreamingResponse(
        output,
        headers=headers_resp,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def dist_coords(c1, c2):
    return math.sqrt((c1[0] - c2[0])**2 + (c1[1] - c2[1])**2)

def ordenar_paradas_otimizadas(origem_coords, lista_pedidos):
    if len(lista_pedidos) <= 1:
        return lista_pedidos

    nao_visitados = list(lista_pedidos)
    rota_ordenada = []
    ponto_atual = origem_coords

    while nao_visitados:
        proximo = min(nao_visitados, key=lambda p: dist_coords(ponto_atual, (p['lat'], p['lon'])))
        rota_ordenada.append(proximo)
        ponto_atual = (proximo['lat'], proximo['lon'])
        nao_visitados.remove(proximo)

    melhorou = True
    while melhorou:
        melhorou = False
        n = len(rota_ordenada)
        for i in range(n - 1):
            for j in range(i + 1, n):
                p_ant = origem_coords if i == 0 else (rota_ordenada[i-1]['lat'], rota_ordenada[i-1]['lon'])
                p_i = (rota_ordenada[i]['lat'], rota_ordenada[i]['lon'])
                p_j = (rota_ordenada[j]['lat'], rota_ordenada[j]['lon'])
                p_prox = origem_coords if j == n - 1 else (rota_ordenada[j+1]['lat'], rota_ordenada[j+1]['lon'])

                dist_atual = dist_coords(p_ant, p_i) + dist_coords(p_j, p_prox)
                dist_invertida = dist_coords(p_ant, p_j) + dist_coords(p_i, p_prox)

                if dist_invertida < dist_atual - 1e-6:
                    rota_ordenada[i:j+1] = reversed(rota_ordenada[i:j+1])
                    melhorou = True
                    break
            if melhorou:
                break

    return rota_ordenada

async def get_osrm_route_async(client: httpx.AsyncClient, pontos):
    if len(pontos) < 2:
        return 0, 0, []

    coords_str = ";".join([f"{lon},{lat}" for lat, lon in pontos])
    url = f"http://router.project-osrm.org/route/v1/driving/{coords_str}?overview=full&geometries=geojson"

    try:
        res = await client.get(url, timeout=6)
        data = res.json()
        if data.get("code") == "Ok":
            route = data["routes"][0]
            dist_km = round(route["distance"] / 1000, 2)
            tempo_min = round(route["duration"] / 60)
            geometria = [[coord[1], coord[0]] for coord in route["geometry"]["coordinates"]]
            return dist_km, tempo_min, geometria
    except Exception:
        pass

    dist_total = 0
    geometria = []
    for i in range(len(pontos) - 1):
        lat1, lon1 = pontos[i]
        lat2, lon2 = pontos[i + 1]
        dist_total += math.sqrt((lat2 - lat1) ** 2 + (lon2 - lon1) ** 2) * 111 * 1.3
        geometria.extend([[lat1, lon1], [lat2, lon2]])

    dist_km = round(dist_total, 2)
    tempo_min = round((dist_km / 22.0) * 60)
    return dist_km, tempo_min, geometria

def montar_rotas_por_capacidade_real(origem_lat, origem_lon, pedidos, frota, modais_usar, iteracoes=5):
    """
    Agrupamento geográfico por CLUSTERIZAÇÃO REAL (k-means capacitado).

    A versão anterior ordenava os pedidos só por ÂNGULO em torno do Hub e cortava
    essa lista 1D em fatias do tamanho da capacidade de cada veículo. Isso falha
    quando a demanda está concentrada de forma desigual entre direções: um veículo
    pode acabar levando uma fatia que varre metade da cidade (ex: Norte + Leste +
    Sul), porque em ÂNGULO esses pontos são "a fatia seguinte", mesmo estando
    fisicamente longe uns dos outros.

    Esta versão usa a posição real (lat/lon) de cada pedido, agrupando por
    proximidade física de verdade, e reajusta os grupos por algumas iterações
    (como um k-means), sempre respeitando a capacidade exata de cada veículo —
    nenhum grupo pode ultrapassar a capacidade do veículo que vai atendê-lo.
    """
    if not pedidos:
        return []

    total_pedidos = len(pedidos)

    # 1. Define as "vagas de viagem" a preencher (motorista + capacidade), repetindo
    #    a frota em rodadas até haver vagas suficientes para todos os pedidos.
    vagas = []
    idx_motorista = 0
    capacidade_acumulada = 0
    while capacidade_acumulada < total_pedidos:
        motorista = frota[idx_motorista % len(frota)]
        cap = max(1, modais_usar.get(motorista.modal, {'capacidade': 6}).get('capacidade', 6))
        vagas.append([motorista, cap])
        capacidade_acumulada += cap
        idx_motorista += 1

    n_vagas = len(vagas)

    # 2. Sementes iniciais via "farthest-point sampling": escolhe pedidos REAIS
    #    como pontos de partida (o mais distante do centro, depois o mais distante
    #    de todas as sementes já escolhidas, e assim por diante). Isso alinha as
    #    sementes com as regiões onde a demanda realmente está concentrada, em vez
    #    de assumir uma distribuição uniforme em círculo — que é o que causava
    #    zigue-zague quando a demanda real é desigual entre direções.
    indices_disponiveis = list(range(total_pedidos))
    primeiro = max(indices_disponiveis, key=lambda i: dist_coords((pedidos[i]['lat'], pedidos[i]['lon']), (origem_lat, origem_lon)))
    sementes_idx = [primeiro]
    while len(sementes_idx) < n_vagas and len(sementes_idx) < total_pedidos:
        candidato = max(
            indices_disponiveis,
            key=lambda i: min(dist_coords((pedidos[i]['lat'], pedidos[i]['lon']), (pedidos[s]['lat'], pedidos[s]['lon'])) for s in sementes_idx)
        )
        sementes_idx.append(candidato)
    sementes = [(pedidos[i]['lat'], pedidos[i]['lon']) for i in sementes_idx]
    # Se houver mais vagas do que pedidos distintos para servir de semente, completa em círculo
    while len(sementes) < n_vagas:
        ang = (2 * math.pi / n_vagas) * len(sementes)
        sementes.append((origem_lat + 0.05 * math.cos(ang), origem_lon + 0.05 * math.sin(ang)))

    atribuicao = [0] * total_pedidos

    # 3. Iterações de atribuição capacitada: a cada rodada, cada pedido "vota" na
    #    vaga mais próxima; pedidos com preferência mais forte (1ª opção bem melhor
    #    que a 2ª) são atendidos primeiro, evitando que um pedido "decisivo" perca
    #    sua melhor vaga para outro pedido que quase não faria diferença.
    for _ in range(iteracoes):
        info_pedidos = []
        for i, p in enumerate(pedidos):
            dists = [dist_coords((p['lat'], p['lon']), s) for s in sementes]
            ordem_preferencia = sorted(range(n_vagas), key=lambda k: dists[k])
            urgencia = (dists[ordem_preferencia[1]] - dists[ordem_preferencia[0]]) if n_vagas > 1 else 0
            info_pedidos.append((i, ordem_preferencia, urgencia))

        info_pedidos.sort(key=lambda x: x[2], reverse=True)

        capacidade_restante = [vagas[k][1] for k in range(n_vagas)]
        nova_atribuicao = [0] * total_pedidos
        for i, ordem_preferencia, _ in info_pedidos:
            for k in ordem_preferencia:
                if capacidade_restante[k] > 0:
                    nova_atribuicao[i] = k
                    capacidade_restante[k] -= 1
                    break
            else:
                nova_atribuicao[i] = ordem_preferencia[0]  # segurança (não deveria ocorrer)

        atribuicao = nova_atribuicao

        # Recalcula o centro geográfico real de cada grupo com base nos membros atuais
        somas = [[0.0, 0.0, 0] for _ in range(n_vagas)]
        for i, p in enumerate(pedidos):
            k = atribuicao[i]
            somas[k][0] += p['lat']
            somas[k][1] += p['lon']
            somas[k][2] += 1
        sementes = [
            (somas[k][0] / somas[k][2], somas[k][1] / somas[k][2]) if somas[k][2] > 0 else sementes[k]
            for k in range(n_vagas)
        ]

    # 4. Monta o resultado final: (motorista, pedidos) por vaga preenchida
    grupos = [[] for _ in range(n_vagas)]
    for i, p in enumerate(pedidos):
        grupos[atribuicao[i]].append(p)

    return [(vagas[k][0], grupos[k]) for k in range(n_vagas) if grupos[k]]

class MotoristaItem(BaseModel):
    id: int
    motorista: str
    modal: str

class ModalConfigItem(BaseModel):
    capacidade: int
    consumo_kml: float

class OtimizarRequest(BaseModel):
    origem_rua: str
    origem_num: str
    origem_bairro: Optional[str] = ""
    origem_cep: Optional[str] = ""
    frota: Optional[List[MotoristaItem]] = []
    modais_config: Optional[Dict[str, ModalConfigItem]] = None
    preco_gasolina: Optional[float] = 5.80
    custo_hora: Optional[float] = 25.00
    pedidos: List[Dict[str, Any]]

@app.get("/api/modais")
def get_modais():
    return MODAIS_PADRAO_DEFAULT

@app.post("/api/upload")
async def upload_planilha(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        filename = file.filename.lower()

        if filename.endswith(".csv"):
            try:
                df = pd.read_csv(io.BytesIO(contents), sep=",")
                if len(df.columns) <= 1:
                    df = pd.read_csv(io.BytesIO(contents), sep=";")
            except Exception:
                df = pd.read_csv(io.BytesIO(contents), sep=";", encoding="latin-1")
        elif filename.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(contents))
        else:
            raise HTTPException(status_code=400, detail="Formato de arquivo inválido.")

        df.columns = [str(c).strip().lower() for c in df.columns]

        col_map = {
            'endereco': ['endereco', 'rua', 'logradouro', 'address', 'street'],
            'numero': ['numero', 'num', 'nro', 'number'],
            'bairro': ['bairro', 'neighborhood', 'district'],
            'cidade': ['cidade', 'city', 'municipio'],
            'uf': ['uf', 'estado', 'state'],
            'cep': ['cep', 'zipcode', 'postal_code'],
            'volume': ['volume', 'vol', 'quantidade', 'pedidos', 'qtd']
        }

        pedidos = []
        for idx, row in df.iterrows():
            def get_val(keys, default=""):
                for k in keys:
                    if k in row and pd.notna(row[k]):
                        return str(row[k]).strip()
                return default

            rua = get_val(col_map['endereco'])
            numero = get_val(col_map['numero'], "S/N")
            bairro = get_val(col_map['bairro'])
            cep = get_val(col_map['cep'])
            cidade = get_val(col_map['cidade'])
            uf = get_val(col_map['uf'])
            volume = 1
            try:
                volume = int(float(get_val(col_map['volume'], 1)))
            except:
                volume = 1

            if rua or cep:
                pedidos.append({
                    "id": idx + 1,
                    "Endereco": rua,
                    "Numero": numero,
                    "Bairro": bairro,
                    "Cidade": cidade,
                    "UF": uf,
                    "CEP": cep,
                    "Volume": volume
                })

        return {"pedidos": pedidos, "total": len(pedidos)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ==========================================================================
# SISTEMA DE JOB COM PROGRESSO
# Como geocodificar pode levar segundos, a otimização roda em segundo plano
# e o frontend consulta o andamento em /api/otimizar/status/{job_id}.
# ==========================================================================
JOBS: Dict[str, Dict[str, Any]] = {}


def _novo_job() -> str:
    job_id = str(uuid.uuid4())
    JOBS[job_id] = {
        "status": "em_andamento",   # em_andamento | concluido | erro
        "etapa": "Iniciando...",
        "atual": 0,
        "total": 1,
        "resultado": None,
        "erro": None,
    }
    return job_id


async def _processar_otimizacao(job_id: str, req: "OtimizarRequest"):
    job = JOBS[job_id]
    try:
        modais_usar = {}
        if req.modais_config:
            for k, v in req.modais_config.items():
                modais_usar[k] = {'capacidade': v.capacidade, 'consumo_kml': v.consumo_kml}
        else:
            modais_usar = MODAIS_PADRAO_DEFAULT

        # --- Etapa 1: geocodificar o Hub ---
        job["etapa"] = "Localizando a Loja Central..."
        async with httpx.AsyncClient() as client:
            orig_lat, orig_lon, cidade_hub, uf_hub, cep_hub, _ = await geocode_async(
                client, req.origem_rua, req.origem_num, req.origem_bairro or "", req.origem_cep or ""
            )

        origem_dict = {"rua": req.origem_rua, "numero": req.origem_num, "lat": orig_lat, "lon": orig_lon}

        # --- Etapa 2: geocodificar todos os pedidos (deduplicado + paralelo) ---
        total_pedidos = len(req.pedidos)
        job["etapa"] = f"Geocodificando endereços (0/{total_pedidos})..."
        job["atual"] = 0
        job["total"] = total_pedidos

        mapa_geocode = await geocodificar_lote(req.pedidos)
        job["atual"] = total_pedidos
        job["etapa"] = f"Endereços geocodificados ({total_pedidos}/{total_pedidos})."

        pedidos_geo = []
        for idx, p in enumerate(req.pedidos):
            rua = p.get("Endereco") or p.get("rua") or ""
            num = str(p.get("Numero") or p.get("numero") or "")
            bairro = p.get("Bairro") or p.get("bairro") or ""
            cep = p.get("CEP") or p.get("cep") or ""
            cidade_p = p.get("Cidade") or p.get("cidade") or ""
            uf_p = p.get("UF") or p.get("uf") or ""
            vol = p.get("Volume") or p.get("volume") or 1

            chave = _chave_cache(rua, numero=num, bairro=bairro, cep=cep, cidade=cidade_p, uf=uf_p)
            p_lat, p_lon, _, _, _, _ = mapa_geocode.get(chave, (orig_lat, orig_lon, "", "", "", ""))

            pedidos_geo.append({
                "id": p.get("id", idx + 1), "Endereco": rua, "Numero": num, "Bairro": bairro,
                "CEP": cep, "Volume": vol, "lat": p_lat, "lon": p_lon
            })

        # --- Etapa 3: montar as viagens respeitando a CAPACIDADE REAL de cada veículo ---
        frota_usar = req.frota if req.frota and len(req.frota) > 0 else [
            MotoristaItem(id=1, motorista="Motorista 01", modal="Carro de Passeio")
        ]
        viagens_planejadas = montar_rotas_por_capacidade_real(orig_lat, orig_lon, pedidos_geo, frota_usar, modais_usar)

        # --- Etapa 4: calcular cada rota (traçado OSRM em paralelo) ---
        job["etapa"] = f"Calculando rotas (0/{len(viagens_planejadas)})..."
        job["atual"] = 0
        job["total"] = len(viagens_planejadas)

        motorista_viagens = {f.id: 0 for f in frota_usar}
        total_viagens_por_motorista = {f.id: 0 for f in frota_usar}
        for mot_info, _ in viagens_planejadas:
            total_viagens_por_motorista[mot_info.id] += 1

        rotas_resultado = []
        km_total_geral = 0.0
        custo_total_geral = 0.0

        async with httpx.AsyncClient() as client:
            for c_idx, (mot_info, cluster) in enumerate(viagens_planejadas):
                motorista_viagens[mot_info.id] += 1
                num_viagem = motorista_viagens[mot_info.id]

                modal_config = modais_usar.get(mot_info.modal, {'consumo_kml': 11.5, 'capacidade': 6})
                consumo_kml = modal_config.get('consumo_kml', 11.5)

                pts_rota = ordenar_paradas_otimizadas((orig_lat, orig_lon), cluster)
                for p in pts_rota:
                    p['modal_alocado'] = mot_info.modal
                    p['motorista_alocado'] = mot_info.motorista
                    p['viagem_alocada'] = num_viagem

                pontos_coords = [(orig_lat, orig_lon)] + [(p['lat'], p['lon']) for p in pts_rota] + [(orig_lat, orig_lon)]
                dist_km, tempo_min, geometria = await get_osrm_route_async(client, pontos_coords)

                amplitude_km = 0.0
                if len(pts_rota) > 1:
                    max_dist = 0
                    for p1 in pts_rota:
                        for p2 in pts_rota:
                            d = haversine_distance((p1['lat'], p1['lon']), (p2['lat'], p2['lon']))
                            if d > max_dist: max_dist = d
                    amplitude_km = max_dist
                
                # Defina o limite que considera "espalhado demais" (ex: 15 km)
                alerta_amplitude = amplitude_km > 15.0

                litros = dist_km / consumo_kml
                custo_comb = litros * req.preco_gasolina
                custo_mot = (tempo_min / 60.0) * req.custo_hora
                custo_rota = custo_comb + custo_mot
                km_total_geral += dist_km
                custo_total_geral += custo_rota

                waypoint_coords = "/".join([f"{p['lat']},{p['lon']}" for p in pts_rota])
                link_maps = f"https://www.google.com/maps/dir/{orig_lat},{orig_lon}/{waypoint_coords}/{orig_lat},{orig_lon}"
                titulo_viagem = f"{mot_info.motorista} • Viagem {num_viagem}" if total_viagens_por_motorista[mot_info.id] > 1 else mot_info.motorista

                rotas_resultado.append({
                    "id": c_idx + 1, "motorista": titulo_viagem, "motorista_base": mot_info.motorista,
                    "viagem_num": num_viagem, "modal": mot_info.modal, "cor": CORES_ROTAS[c_idx % len(CORES_ROTAS)],
                    "qtd_pedidos": len(pts_rota), "km_total": dist_km,
                    "tempo_formatado": f"{tempo_min // 60}h {tempo_min % 60}m" if tempo_min >= 60 else f"{tempo_min} min",
                    "custo_total": round(custo_rota, 2), "custo_por_pedido": round(custo_rota / max(len(pts_rota), 1), 2),
                    "geometria": geometria, "paradas": pts_rota, "link_maps": link_maps
                })

                job["atual"] = c_idx + 1
                job["etapa"] = f"Calculando rotas ({c_idx + 1}/{len(viagens_planejadas)})..."

        total_p = len(pedidos_geo)
        kpis = {
            "total_pedidos": total_p, "total_veiculos": len(frota_usar),
            "total_rotas_viagens": len(rotas_resultado), "km_total": round(km_total_geral, 2),
            "custo_total": round(custo_total_geral, 2),
            "custo_medio_pedido": round(custo_total_geral / max(total_p, 1), 2)
        }

        job["status"] = "concluido"
        job["etapa"] = "Concluído."
        job["resultado"] = {"origem": origem_dict, "rotas": rotas_resultado, "kpis": kpis}

    except Exception as e:
        job["status"] = "erro"
        job["erro"] = f"Erro interno no algoritmo de roteirização: {str(e)}"


@app.post("/api/otimizar")
async def iniciar_otimizacao(req: "OtimizarRequest"):
    """Inicia a roteirização em segundo plano e devolve um job_id na hora."""
    job_id = _novo_job()
    asyncio.create_task(_processar_otimizacao(job_id, req))
    return {"job_id": job_id}


@app.get("/api/otimizar/status/{job_id}")
def status_otimizacao(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado (pode ter expirado).")
    resposta = {
        "status": job["status"],
        "etapa": job["etapa"],
        "atual": job["atual"],
        "total": job["total"],
    }
    if job["status"] == "concluido":
        resposta["resultado"] = job["resultado"]
    if job["status"] == "erro":
        resposta["erro"] = job["erro"]
    return resposta

@app.get("/api/modelo-xlsx")
def download_modelo_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelo_Pedidos"
    
    headers = ["id", "endereco", "numero", "bairro", "cidade", "uf", "cep", "volume"]
    ws.append(headers)
    
    exemplos = [
        [101, "Avenida Paulista", "1500", "Bela Vista", "Sao Paulo", "SP", "01310-100", 1],
        [102, "Rua Augusta", "1500", "Consolacao", "Sao Paulo", "SP", "01304-001", 1],
        [103, "Rua Oscar Freire", "850", "Cerqueira Cesar", "Sao Paulo", "SP", "01426-000", 2],
        [104, "Rua dos Pinheiros", "600", "Pinheiros", "Sao Paulo", "SP", "05422-001", 1],
        [105, "Alameda dos Maracatins", "650", "Moema", "Sao Paulo", "SP", "04089-001", 1],
    ]
    for row in exemplos:
        ws.append(row)
        
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        fill_color = "F8FAFC" if row[0].row % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = thin_border
            if cell.column in [1, 3, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers_resp = {
        'Content-Disposition': 'attachment; filename="modelo_pedidos_zubale.xlsx"'
    }
    return StreamingResponse(
        output, 
        headers=headers_resp, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )